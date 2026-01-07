import io

from decimal import Decimal

from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.forms import formset_factory
from django.http import HttpResponse
from django.db.models import Q, CharField
from django.db.models.functions import Cast

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph

from .forms import SalesTransactionForm, SalesItemForm
from .models import SalesTransaction, SalesItem
from inventory.models import InventoryItem
from storage.services import adjust_storage_for_inventory_item
from storage.models import Packaging
from production.models import ProductPrice
import json
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib import messages
from django.utils import timezone
from production.models import ProductPrice
from customers.models import Customer, LoyaltyTier



class SalesDashboardView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'sales.view_salestransaction'

    def get(self, request):
        queryset = SalesTransaction.objects.select_related('customer').prefetch_related('items__inventory_item').order_by('-created_at')
        queryset = self._apply_filters(queryset, request)
        export = (request.GET.get('export') or '').lower()
        if export == 'pdf':
            return self._export_pdf(queryset)
        if export == 'excel':
            return self._export_excel(queryset)

        # Bulk sales filter: show only transactions with at least one item sold as carton
        show_bulk = request.GET.get('bulk', '') == '1'
        if show_bulk:
            queryset = queryset.filter(items__sold_as='carton').distinct()

        transactions = list(queryset[:20])
        return render(request, 'sales/dashboard.html', {
            'transactions': transactions,
            'show_bulk': show_bulk,
        })

    def _apply_filters(self, queryset, request):
        filters = {
            'transaction_id': (request.GET.get('filter__transaction_id') or '').strip(),
            'customer': (request.GET.get('filter__customer') or '').strip(),
            'total_amount': (request.GET.get('filter__total_amount') or '').strip(),
            'payment_status': (request.GET.get('filter__payment_status') or '').strip(),
            'created_at': (request.GET.get('filter__created_at') or '').strip(),
        }
        if filters['transaction_id']:
            queryset = queryset.filter(transaction_id__icontains=filters['transaction_id'])
        if filters['customer']:
            value = filters['customer']
            queryset = queryset.filter(
                Q(customer__name__icontains=value) |
                Q(walk_in_customer_name__icontains=value) |
                Q(customer_phone__icontains=value)
            )
        if filters['total_amount']:
            queryset = queryset.annotate(_total_str=Cast('total_amount', CharField())).filter(_total_str__icontains=filters['total_amount'])
        if filters['payment_status']:
            queryset = queryset.filter(payment_status__icontains=filters['payment_status'])
        if filters['created_at']:
            queryset = queryset.annotate(_created_str=Cast('created_at', CharField())).filter(_created_str__icontains=filters['created_at'])
        return queryset

    def _export_excel(self, queryset):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Sales'
        headers = ['ID', 'Customer', 'Products', 'Qty', 'Total', 'Status', 'Date']
        sheet.append(headers)
        for tx in queryset:
            products = ', '.join([item.inventory_item.name for item in tx.items.all()])
            quantities = ', '.join([str(int(item.quantity)) for item in tx.items.all()])
            sheet.append([
                tx.transaction_id,
                tx.customer_display_name,
                products or '—',
                quantities or '—',
                float(tx.total_amount or 0),
                tx.get_payment_status_display(),
                timezone.localtime(tx.created_at).strftime('%Y-%m-%d %H:%M'),
            ])
        for idx, _ in enumerate(headers, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = 20
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        timestamp = timezone.now().strftime('%Y%m%d-%H%M%S')
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="sales-{timestamp}.xlsx"'
        return response

    def _export_pdf(self, queryset):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), title='Sales Export')
        styles = getSampleStyleSheet()
        data = [[
            'ID',
            'Customer',
            'Products',
            'Qty',
            'Total',
            'Status',
            'Date',
        ]]
        for tx in queryset:
            products = ', '.join([item.inventory_item.name for item in tx.items.all()])
            quantities = ', '.join([str(int(item.quantity)) for item in tx.items.all()])
            data.append([
                tx.transaction_id,
                tx.customer_display_name,
                products or '—',
                quantities or '—',
                f"KES {tx.total_amount:.2f}",
                tx.get_payment_status_display(),
                timezone.localtime(tx.created_at).strftime('%Y-%m-%d %H:%M'),
            ])
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F2F2F2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#333333')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#DDDDDD')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FBFBFB')]),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        title = Paragraph('Sales Export', styles['Title'])
        doc.build([title, table])
        pdf_value = buffer.getvalue()
        buffer.close()
        timestamp = timezone.now().strftime('%Y%m%d-%H%M%S')
        response = HttpResponse(pdf_value, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="sales-{timestamp}.pdf"'
        return response

class SalesCreateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'sales.add_salestransaction'
    SalesItemFormSet = formset_factory(SalesItemForm, extra=1)

    def get(self, request):
        form = SalesTransactionForm()
        formset = self.SalesItemFormSet()
        # packaging map: {inventory_item_id: packets_per_carton}
        packaging_map = {}
        for pkg in Packaging.objects.select_related('product').all():
            packaging_map[str(pkg.product_id)] = pkg.packets_per_carton

        # product bulk price map: {inventory_item_id: bulk_price_per_carton}
        price_map = {}
        for pp in ProductPrice.objects.select_related('inventory_item').all():
            price_map[str(pp.inventory_item_id)] = str(getattr(pp, 'bulk_price_per_carton', ''))

        return render(request, 'sales/create.html', {
            'form': form,
            'formset': formset,
            'packaging_map_json': json.dumps(packaging_map),
            'product_price_map_json': json.dumps(price_map),
        })

    def _resolve_price(self, inventory_item, request):
        """Return a tuple of (unit_price, bulk_price_per_carton).
        `bulk_price_per_carton` may be None when not available.
        """
        product_price = ProductPrice.current_for_inventory(inventory_item)
        if product_price:
            return product_price.price, getattr(product_price, 'bulk_price_per_carton', None)
        messages.warning(
            request,
            f"No ProductPrice found for {inventory_item.name}. Falling back to inventory default price.",
        )
        return inventory_item.default_price, None

    def post(self, request):
        form = SalesTransactionForm(request.POST)
        formset = self.SalesItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            # Check inventory availability before processing
            insufficient_items = []
            for item_form in formset:
                if item_form.cleaned_data:
                    inventory_item = item_form.cleaned_data['inventory_item']
                    cartons = int(item_form.cleaned_data.get('cartons') or 0)
                    loose = int(item_form.cleaned_data.get('loose_units') or 0)
                    quantity = item_form.cleaned_data.get('quantity')

                    # Determine requested packets
                    total_packets = 0
                    try:
                        pkg = inventory_item.packagings.order_by('-pack_size_ml').first()
                        per_carton = pkg.packets_per_carton if pkg else 1
                    except Exception:
                        per_carton = 1

                    if cartons or loose:
                        total_packets = (cartons * per_carton) + loose
                    elif quantity is not None:
                        total_packets = int(quantity)

                    if total_packets > inventory_item.current_quantity:
                        insufficient_items.append({
                            'name': inventory_item.name,
                            'requested': total_packets,
                            'available': inventory_item.current_quantity
                        })
            
            if insufficient_items:
                for item in insufficient_items:
                    messages.error(
                        request,
                        f"Insufficient stock for {item['name']}: requested {item['requested']}, "
                        f"only {item['available']} available."
                    )
                return render(request, 'sales/create.html', {'form': form, 'formset': formset})
            
            transaction = form.save()
            total = 0
            for item_form in formset:
                if item_form.cleaned_data:
                    inventory_item = item_form.cleaned_data['inventory_item']
                    cartons = int(item_form.cleaned_data.get('cartons') or 0)
                    loose = int(item_form.cleaned_data.get('loose_units') or 0)
                    bulk_price = item_form.cleaned_data.get('bulk_price_per_carton')
                    quantity = item_form.cleaned_data.get('quantity')
                    unit_price, product_bulk_price = self._resolve_price(inventory_item, request)

                    # Determine total packets to deduct from inventory
                    total_packets = 0
                    per_carton = 1
                    try:
                        pkg = inventory_item.packagings.order_by('-pack_size_ml').first()
                        if pkg:
                            per_carton = pkg.packets_per_carton
                    except Exception:
                        pkg = None

                    if cartons or loose:
                        total_packets = (cartons * per_carton) + loose
                    elif quantity is not None:
                        # legacy path: quantity represents packet count
                        total_packets = int(quantity)

                    # Determine effective bulk price with precedence:
                    # manual form override > Packaging.bulk_price_per_carton > ProductPrice.bulk_price_per_carton
                    pkg_bulk = getattr(pkg, 'bulk_price_per_carton', None) if pkg else None
                    effective_bulk = None
                    bulk_source = None
                    if bulk_price is not None:
                        effective_bulk = bulk_price
                        bulk_source = 'manual'
                    elif pkg_bulk:
                        effective_bulk = pkg_bulk
                        bulk_source = 'packaging'
                    elif product_bulk_price:
                        effective_bulk = product_bulk_price
                        bulk_source = 'product'

                    # Enforce business rule: apply bulk price only when the entire
                    # requested quantity is an exact multiple of a carton. Otherwise
                    # sell everything at the unit price.
                    cartons_to_record = 0
                    loose_to_record = 0
                    line_total = Decimal('0')
                    sold_as = 'unit'
                    if per_carton and per_carton > 0 and (total_packets % per_carton) == 0 and effective_bulk:
                        cartons_to_record = total_packets // per_carton
                        loose_to_record = 0
                        line_total = Decimal(cartons_to_record) * Decimal(str(effective_bulk))
                        sold_as = 'carton'
                    else:
                        cartons_to_record = 0
                        loose_to_record = total_packets
                        line_total = Decimal(total_packets) * unit_price
                        sold_as = 'unit'

                    SalesItem.objects.create(
                        transaction=transaction,
                        inventory_item=inventory_item,
                        quantity=Decimal(total_packets),
                        price_per_unit=unit_price,
                        cartons=cartons_to_record,
                        loose_units=loose_to_record,
                        bulk_price_per_carton=(effective_bulk if cartons_to_record else None),
                        sold_as=sold_as,
                        bulk_price_source=bulk_source,
                    )

                    # Deduct stock (packets)
                    from decimal import Decimal as _D
                    inventory_item.current_quantity -= _D(total_packets)
                    inventory_item.save(update_fields=['current_quantity'])
                    adjust_storage_for_inventory_item(inventory_item, -total_packets)

                    total += line_total
            transaction.total_amount = total
            transaction.save(update_fields=['total_amount'])
            if transaction.customer and total:
                points_earned = LoyaltyTier.points_for_amount(total)
                if points_earned:
                    customer = transaction.customer
                    customer.loyalty_points += points_earned
                    customer.save(update_fields=['loyalty_points'])
                    customer.loyalty_ledger.create(
                        points_change=points_earned,
                        balance_after=customer.loyalty_points,
                        reason=f"Purchase {transaction.transaction_id} ({total:.2f})"
                    )
            return redirect('sales:dashboard')
        # On invalid, re-provide maps for the template
        packaging_map = {}
        for pkg in Packaging.objects.select_related('product').all():
            packaging_map[str(pkg.product_id)] = pkg.packets_per_carton
        price_map = {}
        for pp in ProductPrice.objects.select_related('inventory_item').all():
            price_map[str(pp.inventory_item_id)] = str(getattr(pp, 'bulk_price_per_carton', ''))
        return render(request, 'sales/create.html', {'form': form, 'formset': formset, 'packaging_map_json': json.dumps(packaging_map), 'product_price_map_json': json.dumps(price_map)})

class SalesReceiptView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'sales.view_salestransaction'

    def get(self, request, pk):
        transaction = SalesTransaction.objects.prefetch_related('items__inventory_item').get(pk=pk)
        items_qs = transaction.items.select_related('inventory_item').all()
        line_items = []
        subtotal = Decimal('0.00')
        for item in items_qs:
            try:
                pkg = item.inventory_item.packagings.order_by('-pack_size_ml').first()
                per_carton = pkg.packets_per_carton if pkg else 1
            except Exception:
                pkg = None
                per_carton = 1

            cartons = int(getattr(item, 'cartons', 0) or 0)
            loose = int(getattr(item, 'loose_units', 0) or 0)

            if cartons or loose:
                parts = []
                if cartons:
                    parts.append(f"{cartons} carton(s)")
                if loose:
                    parts.append(f"{loose} loose")
                display_qty = ' + '.join(parts)
                total_packets = (cartons * per_carton) + loose
            else:
                display_qty = f"{int(item.quantity)}"
                total_packets = int(item.quantity)

            bulk_applied = bool(item.bulk_price_per_carton is not None and cartons > 0)
            bulk_price = item.bulk_price_per_carton
            line_total = item.line_total
            subtotal += Decimal(line_total)

            # compute per-line discount when bulk applied
            discount = Decimal('0.00')
            if bulk_applied:
                normal_price = (Decimal(total_packets) * item.price_per_unit)
                bulk_total = Decimal(cartons) * Decimal(str(bulk_price))
                discount = normal_price - bulk_total if normal_price > bulk_total else Decimal('0.00')

            line_items.append({
                'inventory_item': item.inventory_item,
                'display_qty': display_qty,
                'price_per_unit': item.price_per_unit,
                'line_total': line_total,
                'cartons': cartons,
                'loose_units': loose,
                'bulk_applied': bulk_applied,
                'bulk_price_per_carton': bulk_price,
                'discount': discount,
                'quantity': item.quantity,
            })

        discount_amount = subtotal - transaction.total_amount if subtotal > transaction.total_amount else Decimal('0.00')
        tax_amount = Decimal('0.00')

        def resolve_user_name(user):
            if not user:
                return None
            full_name = user.get_full_name() if hasattr(user, 'get_full_name') else ''
            return full_name or getattr(user, 'username', None)

        def resolve_job_title(user):
            if not user:
                return None
            profile = getattr(user, 'profile', None)
            job_title = getattr(profile, 'job_title', None) if profile else None
            return job_title

        served_by_user = getattr(transaction, 'created_by', None)
        fallback_user = request.user if request.user.is_authenticated else None
        served_by_name = resolve_user_name(served_by_user) or resolve_user_name(fallback_user) or 'Sales Team'
        served_by_role = resolve_job_title(served_by_user) or resolve_job_title(fallback_user) or 'Sales Clerk'

        return render(request, 'sales/receipt.html', {
            'transaction': transaction,
            'line_items': line_items,
            'subtotal': subtotal,
            'discount_amount': discount_amount,
            'tax_amount': tax_amount,
            'grand_total': transaction.total_amount,
            'served_by_name': served_by_name,
            'served_by_role': served_by_role,
        })


class SalesUpdateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'sales.change_salestransaction'

    def get(self, request, pk):
        transaction = get_object_or_404(SalesTransaction, pk=pk)
        form = SalesTransactionForm(instance=transaction)
        return render(request, 'sales/transaction_form.html', {
            'form': form,
            'transaction': transaction,
            'title': 'Edit Sale',
            'submit_label': 'Update Sale'
        })

    def post(self, request, pk):
        transaction = get_object_or_404(SalesTransaction, pk=pk)
        form = SalesTransactionForm(request.POST, instance=transaction)
        if form.is_valid():
            form.save()
            messages.success(request, f"Sale {transaction.transaction_id} updated successfully.")
            return redirect('sales:dashboard')
        return render(request, 'sales/transaction_form.html', {
            'form': form,
            'transaction': transaction,
            'title': 'Edit Sale',
            'submit_label': 'Update Sale'
        })


class SalesDeleteView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'sales.delete_salestransaction'

    def get(self, request, pk):
        transaction = get_object_or_404(SalesTransaction, pk=pk)
        return render(request, 'sales/transaction_confirm_delete.html', {'transaction': transaction})

    def post(self, request, pk):
        transaction = get_object_or_404(SalesTransaction, pk=pk)
        from decimal import Decimal as _D
        for line in transaction.items.select_related('inventory_item'):
            item = line.inventory_item
            # Restore by computing total packets when cartons/loose_units present
            try:
                per_carton = 1
                pkg = item.packagings.order_by('-pack_size_ml').first()
                if pkg:
                    per_carton = pkg.packets_per_carton
            except Exception:
                pkg = None

            total_restore = None
            if getattr(line, 'cartons', 0) or getattr(line, 'loose_units', 0):
                total_restore = (int(getattr(line, 'cartons', 0)) * per_carton) + int(getattr(line, 'loose_units', 0))
            else:
                total_restore = int(line.quantity)

            item.current_quantity += _D(total_restore)
            item.save(update_fields=['current_quantity'])
            adjust_storage_for_inventory_item(item, total_restore)
        txn_id = transaction.transaction_id
        transaction.delete()
        messages.success(request, f"Sale {txn_id} deleted and stock restored.")
        return redirect('sales:dashboard')
